import asyncio
import json
import os
import random
import re
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional
from urllib.parse import quote_plus

import mcp.server.stdio
import mcp.types as types
from mcp.server import Server
from mcp.server.lowlevel import NotificationOptions
from mcp.server.models import InitializationOptions
from openpyxl import Workbook, load_workbook
from playwright.async_api import BrowserContext, Page, TimeoutError as PlaywrightTimeoutError
from playwright.async_api import async_playwright


APP_NAME = "linkedin-job-apply-agent"
APP_VERSION = "1.1.0"
DEFAULT_EXCEL_FILE = "job_applications.xlsx"
DEFAULT_USER_DATA_DIR = str((Path.home() / ".pw_linkedin_profile").resolve())
DEFAULT_HEADLESS = os.environ.get("LINKEDIN_HEADLESS", "false").lower() == "true"

# FIX 4: Increased slow_mo from 50 to 200 to reduce bot detection
DEFAULT_SLOW_MO_MS = int(os.environ.get("LINKEDIN_SLOWMO_MS", "200"))


# Enhanced anti-detection settings
DEFAULT_ENHANCED_DELAYS = True
DEFAULT_RANDOM_VIEWPORT = True
DEFAULT_RANDOM_USER_AGENT = True

server = Server(APP_NAME)


@dataclass
class JobResult:
    title: str
    company: str
    location: str
    url: str
    status: str = "found"
    notes: str = ""

    def to_dict(self) -> Dict[str, Any]:
        return {
            "title": self.title,
            "company": self.company,
            "location": self.location,
            "url": self.url,
            "status": self.status,
            "notes": self.notes,
        }


def _ok(payload: Dict[str, Any]) -> List[types.TextContent]:
    payload["timestamp"] = datetime.utcnow().isoformat() + "Z"
    return [types.TextContent(type="text", text=json.dumps(payload, ensure_ascii=False))]


def _error(message: str, details: Optional[Dict[str, Any]] = None) -> List[types.TextContent]:
    body: Dict[str, Any] = {"ok": False, "error": message}
    if details:
        body["details"] = details
    return _ok(body)


def _safe_int(value: Any, default: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _normalize_url(url: str) -> str:
    if not url:
        return ""
    url = url.strip()
    if url.startswith("/"):
        url = "https://www.linkedin.com" + url
    if "linkedin.com/jobs/view/" in url:
        return url.split("?")[0]
    return url


def _clean_text(value: Optional[str]) -> str:
    if not value:
        return ""
    return re.sub(r"\s+", " ", value).strip()


def _dedupe_repeated_text(value: str) -> str:
    value = _clean_text(value)
    if not value:
        return value
    if len(value) % 2 == 0:
        half = len(value) // 2
        if value[:half] == value[half:]:
            return value[:half]
    return value


def _load_resume_text(resume_path: str) -> str:
    path = Path(resume_path).expanduser().resolve()
    if not path.exists():
        return ""

    suffix = path.suffix.lower()
    try:
        if suffix in {".txt", ".md"}:
            return path.read_text(encoding="utf-8", errors="ignore")
        if suffix == ".pdf":
            try:
                from pypdf import PdfReader
                reader = PdfReader(str(path))
                return "\n".join((page.extract_text() or "") for page in reader.pages)
            except Exception:
                return ""
        if suffix == ".docx":
            try:
                import docx
                doc = docx.Document(str(path))
                return "\n".join(p.text for p in doc.paragraphs)
            except Exception:
                return ""
    except Exception:
        return ""
    return ""


def _extract_first(pattern: str, text: str) -> str:
    match = re.search(pattern, text, flags=re.IGNORECASE)
    return _clean_text(match.group(1) if match else "")


def _build_candidate_profile(
    phone: str = "",
    cover_letter: str = "",
    resume_path: str = "",
    candidate_profile: Optional[Dict[str, Any]] = None,
) -> Dict[str, str]:
    profile: Dict[str, str] = {}
    if candidate_profile:
        for key, value in candidate_profile.items():
            if value is not None:
                profile[str(key).lower()] = _clean_text(str(value))

    if phone:
        profile["phone"] = _clean_text(phone)
    if cover_letter:
        profile["cover_letter"] = _clean_text(cover_letter)

    if resume_path:
        text = _load_resume_text(resume_path)
        if text:
            email = _extract_first(r"([A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})", text)
            phone_match = _extract_first(r"(\+?\d[\d\-\s()]{8,}\d)", text)
            years = _extract_first(r"(\d+)\+?\s+years", text)
            if email and "email" not in profile:
                profile["email"] = email
            if phone_match and "phone" not in profile:
                profile["phone"] = phone_match
            if years and "years_experience" not in profile:
                profile["years_experience"] = years

    return profile


async def _human_pause(min_s: float = 0.8, max_s: float = 1.8) -> None:
    # FIX 4: Slightly longer pauses to reduce bot detection
    await asyncio.sleep(random.uniform(min_s, max_s))


def _get_random_user_agent() -> str:
    """Generate a random Chrome user agent to avoid detection."""
    chrome_versions = [
        "124.0.0.0", "123.0.0.0", "122.0.0.0", "121.0.0.0", "120.0.0.0",
        "119.0.0.0", "118.0.0.0", "117.0.0.0", "116.0.0.0", "115.0.0.0"
    ]
    chrome_version = random.choice(chrome_versions)
    windows_versions = [
        "Windows NT 10.0; Win64; x64",
        "Windows NT 10.0",
        "Windows NT 6.3; Win64; x64",
        "Windows NT 6.1; Win64; x64"
    ]
    windows_version = random.choice(windows_versions)
    
    return f"Mozilla/5.0 ({windows_version}) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/{chrome_version} Safari/537.36"


def _get_random_viewport() -> Dict[str, int]:
    """Generate a random viewport size to appear more human-like."""
    viewports = [
        {"width": 1920, "height": 1080},
        {"width": 1366, "height": 768},
        {"width": 1536, "height": 864},
        {"width": 1440, "height": 900},
        {"width": 1280, "height": 720},
        {"width": 1600, "height": 900},
        {"width": 1280, "height": 800},
        {"width": 1024, "height": 768},
    ]
    return random.choice(viewports)


async def _random_mouse_movement(page: Page) -> None:
    """Simulate random mouse movements to appear more human-like."""
    try:
        # Get viewport dimensions
        viewport = page.viewport_size
        if not viewport:
            return
            
        width, height = viewport["width"], viewport["height"]
        
        # Make 3-5 random mouse movements
        for _ in range(random.randint(3, 5)):
            x = random.randint(100, width - 100)
            y = random.randint(100, height - 100)
            
            # Move mouse to random position with human-like speed
            await page.mouse.move(x, y)
            await asyncio.sleep(random.uniform(0.1, 0.3))
            
            # Occasionally click to simulate interaction
            if random.random() < 0.3:
                await page.mouse.click(x, y)
                await asyncio.sleep(random.uniform(0.2, 0.5))
    except Exception:
        pass  # Ignore mouse movement errors


async def _enhanced_delays() -> None:
    """Enhanced human-like delays with more variation."""
    if DEFAULT_ENHANCED_DELAYS:
        # Base delay with more variation
        base_delay = random.uniform(1.5, 3.5)
        # Add random micro-delays
        micro_delays = [random.uniform(0.1, 0.5) for _ in range(random.randint(2, 4))]
        
        await asyncio.sleep(base_delay)
        for delay in micro_delays:
            await asyncio.sleep(delay)
    else:
        await _human_pause(0.8, 1.8)


async def _launch_context(user_data_dir: Optional[str] = None) -> BrowserContext:
    playwright = await async_playwright().start()
    profile_dir = user_data_dir or DEFAULT_USER_DATA_DIR
    
    # Enhanced anti-detection: Use randomized settings
    viewport = _get_random_viewport() if DEFAULT_RANDOM_VIEWPORT else {"width": 1440, "height": 900}
    user_agent = _get_random_user_agent() if DEFAULT_RANDOM_USER_AGENT else (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
    
    context = await playwright.chromium.launch_persistent_context(
        user_data_dir=profile_dir,
        channel="chrome",
        headless=DEFAULT_HEADLESS,
        slow_mo=DEFAULT_SLOW_MO_MS,
        viewport=viewport,
        user_agent=user_agent,
    )
    context._linked_playwright = playwright
    
    # Enhanced anti-detection: Add random mouse movements and delays after launch
    if context.pages:
        page = context.pages[0]
        await _random_mouse_movement(page)
        await _enhanced_delays()
    
    return context


async def _close_context(context: BrowserContext) -> None:
    playwright = getattr(context, "_linked_playwright", None)
    await context.close()
    if playwright:
        await playwright.stop()


async def _ensure_logged_in(page: Page, navigate: bool = True) -> bool:
    if navigate:
        await page.goto("https://www.linkedin.com/feed/", wait_until="domcontentloaded")
        await _human_pause()
    if "linkedin.com/login" in page.url.lower() or "checkpoint" in page.url.lower():
        return False
    return True


async def _linkedin_search(
    page: Page,
    role: str,
    location: str,
    count: int,
) -> List[JobResult]:
    keywords = quote_plus(role)
    geo = quote_plus(location)
    # Added f_TPR=r86400 to filter jobs posted in last 24 hours
    url = (
        f"https://www.linkedin.com/jobs/search/"
        f"?keywords={keywords}&location={geo}&f_AL=true&f_TPR=r86400"
    )

    await page.goto(url, wait_until="domcontentloaded")
    await page.wait_for_timeout(2000)

    try:
        await page.wait_for_selector(
            "li.scaffold-layout__list-item, div.job-search-card, ul.jobs-search__results-list li",
            timeout=20000,
        )
    except PlaywrightTimeoutError:
        return []

    cards_seen = set()
    jobs: List[JobResult] = []
    no_growth_loops = 0

    for _ in range(40):
        cards = page.locator(
            "li.scaffold-layout__list-item, ul.jobs-search__results-list li, div.job-search-card"
        )
        card_count = await cards.count()

        for idx in range(card_count):
            card = cards.nth(idx)
            link = card.locator("a[href*='/jobs/view/']").first
            if not await link.count():
                continue

            card_text = _clean_text(await card.text_content()).lower()
            if "easy apply" not in card_text:
                continue

            href = await link.get_attribute("href")
            job_url = _normalize_url(href or "")
            title = _dedupe_repeated_text(await link.text_content() or "")

            company = ""
            for company_selector in [
                "div.artdeco-entity-lockup__subtitle",
                "h4.base-search-card__subtitle",
                ".job-card-container__company-name",
                ".base-search-card__subtitle",
            ]:
                node = card.locator(company_selector).first
                if await node.count():
                    company = _clean_text(await node.text_content())
                    if company:
                        break

            place = ""
            for place_selector in [
                "li.job-card-container__metadata-item",
                ".job-search-card__location",
                ".base-search-card__metadata",
            ]:
                node = card.locator(place_selector).first
                if await node.count():
                    place = _clean_text(await node.text_content())
                    if place:
                        break

            if not job_url or job_url in cards_seen:
                continue

            cards_seen.add(job_url)
            jobs.append(
                JobResult(
                    title=title or "Unknown Title",
                    company=company or "Unknown Company",
                    location=place or "Unknown Location",
                    url=job_url,
                )
            )
            if len(jobs) >= count:
                return jobs

        previous_total = len(jobs)
        await page.mouse.wheel(0, random.randint(1800, 2600))
        await _human_pause(0.8, 1.6)

        if len(jobs) == previous_total:
            no_growth_loops += 1
        else:
            no_growth_loops = 0

        if no_growth_loops >= 4:
            break

    return jobs


async def _open_job(page: Page, job_url: str) -> None:
    await page.goto(job_url, wait_until="domcontentloaded")
    await page.wait_for_timeout(1500)


async def _fill_cover_letter_if_possible(page: Page, cover_letter: str) -> bool:
    if not cover_letter:
        return False
    textareas = [
        "textarea[name='coverLetter']",
        "textarea[aria-label*='Cover letter']",
        "textarea[id*='cover']",
        "textarea",
    ]
    for selector in textareas:
        area = page.locator(selector).first
        if await area.count() and await area.is_visible():
            await area.fill(cover_letter)
            return True
    return False


async def _fill_phone_if_possible(page: Page, phone: str) -> bool:
    if not phone:
        return False
    selectors = [
        "input[autocomplete='tel']",
        "input[id*='phone']",
        "input[aria-label*='Phone']",
        "input[name*='phone']",
    ]
    for selector in selectors:
        field = page.locator(selector).first
        if await field.count() and await field.is_visible():
            await field.fill(phone)
            return True
    return False




async def _fill_input_if_context_matches(
    page: Page, context_keys: List[str], value: str, input_selectors: List[str]
) -> int:
    if not value:
        return 0
    filled = 0
    for selector in input_selectors:
        fields = page.locator(selector)
        count = await fields.count()
        for idx in range(min(count, 60)):
            field = fields.nth(idx)
            if not await field.is_visible() or not await field.is_enabled():
                continue
            field_id = _clean_text((await field.get_attribute("id")) or "").lower()
            field_name = _clean_text((await field.get_attribute("name")) or "").lower()
            placeholder = _clean_text((await field.get_attribute("placeholder")) or "").lower()
            aria = _clean_text((await field.get_attribute("aria-label")) or "").lower()
            context = f"{field_id} {field_name} {placeholder} {aria}"
            if any(k in context for k in context_keys):
                try:
                    await field.fill(value)
                    filled += 1
                except Exception:
                    continue
    return filled


async def _fill_select_if_context_matches(
    page: Page, context_keys: List[str], preferred_values: List[str]
) -> int:
    selected = 0
    selects = page.locator("select")
    count = await selects.count()
    for idx in range(min(count, 40)):
        sel = selects.nth(idx)
        if not await sel.is_visible() or not await sel.is_enabled():
            continue
        sel_id = _clean_text((await sel.get_attribute("id")) or "").lower()
        sel_name = _clean_text((await sel.get_attribute("name")) or "").lower()
        aria = _clean_text((await sel.get_attribute("aria-label")) or "").lower()
        context = f"{sel_id} {sel_name} {aria}"
        if not any(k in context for k in context_keys):
            continue
        options = sel.locator("option")
        opt_count = await options.count()
        picked = False
        for p in preferred_values:
            for j in range(opt_count):
                opt = options.nth(j)
                label = _clean_text(await opt.text_content()).lower()
                value = _clean_text((await opt.get_attribute("value")) or "").lower()
                if p in label or p == value:
                    try:
                        await sel.select_option(index=j)
                        selected += 1
                        picked = True
                        break
                    except Exception:
                        pass
            if picked:
                break
    return selected


async def _fill_textarea_if_context_matches(
    page: Page, context_keys: List[str], value: str, textarea_selectors: List[str]
) -> int:
    if not value:
        return 0
    filled = 0
    for selector in textarea_selectors:
        fields = page.locator(selector)
        count = await fields.count()
        for idx in range(min(count, 60)):
            field = fields.nth(idx)
            if not await field.is_visible() or not await field.is_enabled():
                continue
            field_id = _clean_text((await field.get_attribute("id")) or "").lower()
            field_name = _clean_text((await field.get_attribute("name")) or "").lower()
            placeholder = _clean_text((await field.get_attribute("placeholder")) or "").lower()
            aria = _clean_text((await field.get_attribute("aria-label")) or "").lower()
            context = f"{field_id} {field_name} {placeholder} {aria}"
            if any(k in context for k in context_keys):
                try:
                    await field.fill(value)
                    filled += 1
                except Exception:
                    continue
    return filled


async def _fill_checkbox_if_context_matches(
    page: Page, context_keys: List[str], preferred_values: List[str]
) -> int:
    checked = 0
    checkboxes = page.locator("input[type='checkbox']")
    count = await checkboxes.count()
    for idx in range(min(count, 40)):
        checkbox = checkboxes.nth(idx)
        if not await checkbox.is_visible() or not await checkbox.is_enabled():
            continue
        checkbox_id = _clean_text((await checkbox.get_attribute("id")) or "").lower()
        checkbox_name = _clean_text((await checkbox.get_attribute("name")) or "").lower()
        aria = _clean_text((await checkbox.get_attribute("aria-label")) or "").lower()
        context = f"{checkbox_id} {checkbox_name} {aria}"
        if not any(k in context for k in context_keys):
            continue

        label_text = ""
        label = page.locator(f"label[for='{checkbox_id}']").first
        if await label.count():
            label_text = _clean_text(await label.text_content()).lower()

        combined_context = f"{context} {label_text}"
        should_check = any(val in combined_context for val in preferred_values)

        if should_check and not await checkbox.is_checked():
            try:
                await checkbox.check()
                checked += 1
            except Exception:
                pass
    return checked


async def _fill_radio_if_context_matches(
    page: Page, context_keys: List[str], preferred_values: List[str]
) -> int:
    selected = 0
    radios = page.locator("input[type='radio']")
    count = await radios.count()
    for idx in range(min(count, 40)):
        radio = radios.nth(idx)
        if not await radio.is_visible() or not await radio.is_enabled():
            continue
        radio_id = _clean_text((await radio.get_attribute("id")) or "").lower()
        radio_name = _clean_text((await radio.get_attribute("name")) or "").lower()
        aria = _clean_text((await radio.get_attribute("aria-label")) or "").lower()
        context = f"{radio_id} {radio_name} {aria}"
        if not any(k in context for k in context_keys):
            continue

        label_text = ""
        label = page.locator(f"label[for='{radio_id}']").first
        if await label.count():
            label_text = _clean_text(await label.text_content()).lower()

        combined_context = f"{context} {label_text}"
        should_select = any(val in combined_context for val in preferred_values)

        if should_select:
            try:
                await radio.check()
                selected += 1
                break
            except Exception:
                pass
    return selected


async def _fill_file_upload_if_context_matches(
    page: Page, context_keys: List[str], file_path: str
) -> int:
    if not file_path:
        return 0

    uploaded = 0
    file_inputs = page.locator("input[type='file']")
    count = await file_inputs.count()
    for idx in range(min(count, 20)):
        file_input = file_inputs.nth(idx)
        file_id = _clean_text((await file_input.get_attribute("id")) or "").lower()
        file_name = _clean_text((await file_input.get_attribute("name")) or "").lower()
        aria = _clean_text((await file_input.get_attribute("aria-label")) or "").lower()
        context = f"{file_id} {file_name} {aria}"
        if any(k in context for k in context_keys):
            try:
                await file_input.set_input_files(file_path)
                uploaded += 1
            except Exception:
                continue
    return uploaded


async def _enhanced_auto_fill_application_fields(
    page: Page, profile: Dict[str, str], resume_path: str
) -> int:
    updates = 0

    updates += await _fill_input_if_context_matches(
        page, ["name", "full name", "candidate name"],
        profile.get("full_name", profile.get("name", "")),
        ["input[type='text']"],
    )
    updates += await _fill_input_if_context_matches(
        page, ["email", "e-mail"],
        profile.get("email", ""),
        ["input[type='email']", "input[type='text']"],
    )
    updates += await _fill_input_if_context_matches(
        page, ["phone", "mobile", "tel"],
        profile.get("phone", ""),
        ["input[type='tel']", "input[type='text']"],
    )
    updates += await _fill_input_if_context_matches(
        page, ["experience", "years", "yoe"],
        profile.get("years_experience", profile.get("experience_years", "")),
        ["input[type='number']", "input[type='text']"],
    )
    updates += await _fill_input_if_context_matches(
        page, ["notice"],
        profile.get("notice_period", profile.get("notice", "")),
        ["input[type='text']", "input[type='number']"],
    )
    updates += await _fill_input_if_context_matches(
        page, ["salary", "ctc", "compensation", "expected"],
        profile.get("expected_salary", profile.get("salary", "")),
        ["input[type='text']", "input[type='number']"],
    )
    updates += await _fill_input_if_context_matches(
        page, ["current ctc", "current salary", "current compensation", "present ctc", "present salary"],
        profile.get("current_ctc", ""),
        ["input[type='text']", "input[type='number']"],
    )
    updates += await _fill_input_if_context_matches(
        page, ["authorization", "authorized", "sponsorship", "sponsor", "visa"],
        profile.get("work_authorization", ""),
        ["input[type='text']"],
    )
    updates += await _fill_input_if_context_matches(
        page, ["location", "current city", "current location", "city"],
        profile.get("current_location", profile.get("location", "")),
        ["input[type='text']"],
    )
    updates += await _fill_input_if_context_matches(
        page, ["linkedin", "linkedin url", "linkedin profile"],
        profile.get("linkedin_url", ""),
        ["input[type='url']", "input[type='text']"],
    )
    updates += await _fill_input_if_context_matches(
        page, ["github", "github url", "github profile"],
        profile.get("github_url", ""),
        ["input[type='url']", "input[type='text']"],
    )
    updates += await _fill_input_if_context_matches(
        page, ["portfolio", "website", "personal site"],
        profile.get("portfolio_url", ""),
        ["input[type='url']", "input[type='text']"],
    )
    updates += await _fill_input_if_context_matches(
        page, ["date of birth", "dob", "birth"],
        profile.get("dob", ""),
        ["input[type='date']", "input[type='text']"],
    )
    updates += await _fill_input_if_context_matches(
        page, ["graduation", "graduation year", "passing year", "year of graduation"],
        profile.get("graduation_year", ""),
        ["input[type='number']", "input[type='text']"],
    )

    updates += await _fill_textarea_if_context_matches(
        page,
        ["cover", "letter", "summary", "description", "additional", "details"],
        profile.get("cover_letter", ""),
        ["textarea"],
    )

    updates += await _fill_select_if_context_matches(
        page, ["authorization", "authorized", "sponsorship", "sponsor", "visa"],
        ["yes", "no sponsorship", "authorized", "citizen", "not require"],
    )
    updates += await _fill_select_if_context_matches(
        page, ["relocation", "relocate"],
        ["yes", "open", "willing"],
    )
    updates += await _fill_select_if_context_matches(
        page, ["notice"],
        ["immediately", "15", "30", "1 month"],
    )
    updates += await _fill_select_if_context_matches(
        page, ["degree", "education", "qualification"],
        ["bachelor", "master", "phd", "associate"],
    )
    updates += await _fill_select_if_context_matches(
        page, ["availability", "start", "begin"],
        ["immediately", "2 weeks", "1 month", "negotiable"],
    )
    if profile.get("gender"):
        updates += await _fill_select_if_context_matches(
            page, ["gender", "sex"],
            [profile["gender"].strip().lower()],
        )
    if profile.get("willing_to_relocate", "").strip().lower() in {"yes", "true", "y"}:
        updates += await _fill_select_if_context_matches(
            page, ["relocation", "relocate"],
            ["yes", "open", "willing"],
        )
    if profile.get("employment_type_preference"):
        updates += await _fill_select_if_context_matches(
            page, ["employment", "employment type", "work type"],
            [profile["employment_type_preference"].strip().lower()],
        )

    updates += await _fill_checkbox_if_context_matches(
        page, ["agree", "terms", "conditions", "privacy", "consent"],
        ["yes", "agree", "accept"],
    )

    updates += await _fill_radio_if_context_matches(
        page, ["employment", "status", "type"],
        ["full-time", "part-time", "contract", "internship"],
    )
    updates += await _fill_radio_if_context_matches(
        page, ["remote", "work", "location"],
        ["remote", "hybrid", "onsite"],
    )

    if resume_path:
        updates += await _fill_file_upload_if_context_matches(
            page, ["resume", "cv", "portfolio", "document"],
            resume_path,
        )

    return updates


# FIX 5: Check for unfilled required fields before clicking Next
async def _check_required_fields_filled(page: Page) -> bool:
    """Returns True if all required fields appear to be filled."""
    required_inputs = page.locator(
        "input[required]:not([type='hidden']), "
        "select[required], "
        "textarea[required], "
        "[aria-required='true']"
    )
    count = await required_inputs.count()
    for idx in range(min(count, 30)):
        field = required_inputs.nth(idx)
        if not await field.is_visible():
            continue
        value = _clean_text((await field.get_attribute("value")) or "")
        inner = _clean_text(await field.inner_text() if await field.count() else "")
        if not value and not inner:
            # There is a visible required field that appears empty
            return False
    return True


async def _dismiss_save_dialog_if_present(page: Page) -> bool:
    """
    LinkedIn shows 'Save this application?' whenever the Easy Apply dialog
    is interrupted. We always click 'Discard' to clear it so automation
    can continue or re-open the dialog cleanly.
    Returns True if the dialog was found and dismissed.
    """
    try:
        discard_selectors = [
            "button:has-text('Discard')",
            "button[data-test-dialog-secondary-btn]",
            "button.artdeco-modal__confirm-dialog-btn--secondary",
        ]
        for selector in discard_selectors:
            discard_btn = page.locator(selector).first
            if await discard_btn.count() and await discard_btn.is_visible():
                await discard_btn.click()
                await page.wait_for_timeout(800)
                print("Dismissed 'Save this application?' popup.")
                return True
    except Exception:
        pass
    return False


async def _wait_and_dismiss_save_popup(page: Page, timeout_ms: int = 3000) -> None:
    """Wait briefly for a 'Save this application?' popup and dismiss it if it appears."""
    try:
        await page.wait_for_selector("button:has-text('Discard')", timeout=timeout_ms)
        await _dismiss_save_dialog_if_present(page)
    except PlaywrightTimeoutError:
        pass  # No popup, continue normally


async def _smart_form_navigation(page: Page) -> bool:
    """Smart navigation through multi-step forms with required field check."""
    try:
        # Dismiss any stray 'Save this application?' popup before trying to navigate
        await _dismiss_save_dialog_if_present(page)

        # FIX 5: warn if required fields aren't filled, but still attempt to proceed
        all_filled = await _check_required_fields_filled(page)
        if not all_filled:
            print("Warning: Some required fields may not be filled on this step.")

        next_selectors = [
            "button[aria-label='Continue to next step']",
            "button[aria-label='Next']",
            "button:has-text('Next')",
            "button:has-text('Continue')",
            "button:has-text('Continue Application')",
            "button[data-test-id='next-button']",
            "button[data-test-id='continue-button']",
            "button[type='submit']",
            "input[type='submit']",
        ]

        for selector in next_selectors:
            button = page.locator(selector).first
            if await button.count() and await button.is_visible():
                await button.scroll_into_view_if_needed()
                await button.click()
                await page.wait_for_timeout(800)
                # After clicking Next, dismiss any save dialog that pops up
                await _dismiss_save_dialog_if_present(page)
                return True

        submit_selectors = [
            "button[aria-label='Review your application']",
            "button[aria-label='Submit application']",
            "button:has-text('Submit')",
            "button:has-text('Review')",
            "button[data-test-id='submit-button']",
        ]

        for selector in submit_selectors:
            button = page.locator(selector).first
            if await button.count() and await button.is_visible():
                return True

        return False
    except Exception:
        return False


async def _click_if_visible(page: Page, selector: str) -> bool:
    button = page.locator(selector).first
    if await button.count() and await button.is_visible():
        await button.click()
        return True
    return False


async def _is_easy_apply_dialog_open(page: Page) -> bool:
    selectors = [
        "div.jobs-easy-apply-modal",
        "div[role='dialog'] button[aria-label='Submit application']",
        "div[role='dialog'] button[aria-label='Continue to next step']",
        "div[role='dialog'] button[aria-label='Review your application']",
        "div[role='dialog'] h2:has-text('Apply')",
        "div[role='dialog'] h3:has-text('Apply')",
    ]
    for selector in selectors:
        node = page.locator(selector).first
        try:
            if await node.count() and await node.is_visible():
                return True
        except Exception:
            continue
    return False


async def _click_easy_apply(page: Page) -> bool:
    candidates = [
        "button.jobs-apply-button",
        "button[aria-label*='Easy Apply']",
        "button:has-text('Easy Apply')",
        "div.jobs-apply-button--top-card button",
        "button[data-live-test-job-apply-button]",
        "button[aria-label*='easy apply' i]",
        "button:has-text('Easy apply')",
        "a:has-text('Easy Apply')",
        "[role='button']:has-text('Easy Apply')",
    ]
    await page.mouse.wheel(0, 500)
    await page.wait_for_timeout(400)
    for selector in candidates:
        button = page.locator(selector).first
        if await button.count() and await button.is_visible():
            try:
                await button.scroll_into_view_if_needed()
                await button.click(timeout=3000)
            except Exception:
                try:
                    await button.click(force=True, timeout=3000)
                except Exception:
                    continue
            await page.wait_for_timeout(1000)
            if await _is_easy_apply_dialog_open(page):
                return True

    clickables = page.locator("button, a, [role='button']")
    count = await clickables.count()
    for idx in range(min(count, 120)):
        btn = clickables.nth(idx)
        if not await btn.is_visible():
            continue
        text = _clean_text(await btn.text_content())
        aria = _clean_text(await btn.get_attribute("aria-label"))
        combined = f"{text} {aria}".lower()
        if "easy apply" in combined:
            try:
                await btn.scroll_into_view_if_needed()
                await btn.click(timeout=3000)
            except Exception:
                try:
                    await btn.click(force=True, timeout=3000)
                except Exception:
                    continue
            await page.wait_for_timeout(1000)
            if await _is_easy_apply_dialog_open(page):
                return True
    return await _is_easy_apply_dialog_open(page)


async def _collect_visible_button_labels(page: Page, limit: int = 12) -> List[str]:
    labels: List[str] = []
    buttons = page.locator("button, a, [role='button']")
    count = await buttons.count()
    for idx in range(min(count, 140)):
        if len(labels) >= limit:
            break
        btn = buttons.nth(idx)
        if not await btn.is_visible():
            continue
        text = _clean_text(await btn.text_content())
        aria = _clean_text(await btn.get_attribute("aria-label"))
        label = _clean_text(f"{text} {aria}")
        if label:
            labels.append(label)
    return labels


async def _detect_apply_state(page: Page) -> Dict[str, Any]:
    page_text = _clean_text((await page.locator("body").first.text_content()) or "").lower()

    explicit_applied_selectors = [
        "button:has-text('Applied')",
        "span:has-text('Applied')",
        "div:has-text(\"You've applied\")",
        "li:has-text(\"Application submitted\")",
    ]
    for selector in explicit_applied_selectors:
        node = page.locator(selector).first
        if await node.count() and await node.is_visible():
            return {
                "state": "already_applied",
                "message": "Already applied to this job.",
                "visible_buttons": await _collect_visible_button_labels(page),
            }

    if "you've applied" in page_text:
        return {
            "state": "already_applied",
            "message": "Already applied to this job.",
            "visible_buttons": await _collect_visible_button_labels(page),
        }

    if "easy apply" in page_text:
        return {
            "state": "easy_apply_visible_but_not_clickable",
            "message": "Easy Apply text found but button could not be clicked.",
            "visible_buttons": await _collect_visible_button_labels(page),
        }

    if (
        "this job is no longer accepting applications" in page_text
        or "no longer accepting applications" in page_text
    ):
        return {"state": "job_closed", "message": "This job is no longer accepting applications."}

    external_apply = page.locator(
        "a[data-live-test-job-apply-button], a.jobs-apply-button, button:has-text('Apply'), a:has-text('Apply')"
    ).first
    if await external_apply.count() and await external_apply.is_visible():
        label = _clean_text(await external_apply.text_content())
        if "easy apply" not in label.lower():
            return {
                "state": "external_apply",
                "message": "This listing uses external apply, not LinkedIn Easy Apply.",
                "apply_label": label or "Apply",
                "visible_buttons": await _collect_visible_button_labels(page),
            }

    if "verify" in page_text and "human" in page_text:
        return {
            "state": "challenge",
            "message": "LinkedIn is showing a verification/CAPTCHA challenge.",
            "visible_buttons": await _collect_visible_button_labels(page),
        }

    visible_buttons = await _collect_visible_button_labels(page)
    if any("easy apply" in label.lower() for label in visible_buttons):
        return {
            "state": "easy_apply_visible_but_not_clickable",
            "message": "Easy Apply is visible but could not be clicked by automation.",
            "visible_buttons": visible_buttons,
        }

    return {
        "state": "unknown",
        "message": "Easy Apply button not detected.",
        "visible_buttons": visible_buttons,
    }


async def _debug_job_page(page: Page, job_url: str) -> Dict[str, Any]:
    await _open_job(page, job_url)
    visible_buttons = await _collect_visible_button_labels(page, limit=25)
    has_easy_apply_text = any("easy apply" in label.lower() for label in visible_buttons)

    apply_links = []
    links = page.locator("a")
    link_count = await links.count()
    for idx in range(min(link_count, 80)):
        link = links.nth(idx)
        if not await link.is_visible():
            continue
        text = _clean_text(await link.text_content())
        if not text:
            continue
        if "apply" in text.lower():
            href = _normalize_url((await link.get_attribute("href")) or "")
            apply_links.append({"text": text, "href": href})
        if len(apply_links) >= 10:
            break

    return {
        "ok": True,
        "tool": "debug_job_page",
        "job_url": job_url,
        "page_url": page.url,
        "title": await page.title(),
        "has_easy_apply_text": has_easy_apply_text,
        "visible_buttons": visible_buttons,
        "apply_links": apply_links,
    }


async def _run_easy_apply_flow(
    page: Page,
    job_url: str,
    profile: Dict[str, str],
    phone: str,
    cover_letter: str,
    dry_run: bool,
    resume_path: str,
    max_steps: int = 15,
    enhanced: bool = False,
) -> Dict[str, Any]:
    """Unified apply flow used by both easy_apply and enhanced_easy_apply."""

    await _open_job(page, job_url)
    try:
        await page.wait_for_selector(
            "button.jobs-apply-button, button[aria-label*='Easy Apply'], button:has-text('Easy Apply')",
            timeout=8000,
        )
    except PlaywrightTimeoutError:
        pass

    easy_apply_clicked = await _click_easy_apply(page)
    if not easy_apply_clicked:
        diagnostics = await _detect_apply_state(page)
        return {
            "ok": False,
            "job_url": job_url,
            "status": "not_easy_apply",
            "message": "Easy Apply button was not found.",
            "diagnostics": diagnostics,
        }

    # KEY FIX: Right after Easy Apply opens, wait for the
    # "Save this application?" popup and immediately discard it.
    # This popup appears because a previous session left the dialog open.
    await page.wait_for_timeout(1200)
    await _wait_and_dismiss_save_popup(page, timeout_ms=3000)

    # Now fill the first page fields
    await _fill_phone_if_possible(page, profile.get("phone", phone))
    await _fill_cover_letter_if_possible(page, profile.get("cover_letter", cover_letter))
    total_updates = await _enhanced_auto_fill_application_fields(page, profile, resume_path)

    progressed = False
    stall_count = 0

    for step in range(1, max_steps + 1):
        await _human_pause(0.8, 1.4)

        # Clear any popup at the start of every step
        if await _dismiss_save_dialog_if_present(page):
            print(f"Step {step}: Cleared save popup, retrying...")
            await page.wait_for_timeout(800)
            stall_count = 0
            continue

        # --- Check if we've reached the final submit button ---
        submitted = page.locator("button[aria-label='Submit application']").first
        if await submitted.count() and await submitted.is_visible():
            if dry_run:
                return {
                    "ok": True,
                    "job_url": job_url,
                    "status": "dry_run_ready_to_submit",
                    "steps": step,
                    "field_updates": total_updates,
                    "message": (
                        f"Reached final submit step ({total_updates} fields filled). "
                        "Dry run — not submitted. Set dry_run=false to submit for real."
                    ),
                }
            await submitted.click()
            await page.wait_for_timeout(1500)
            await _click_if_visible(page, "button[aria-label='Dismiss']")
            return {
                "ok": True,
                "job_url": job_url,
                "status": "submitted",
                "steps": step,
                "field_updates": total_updates,
                "message": f"Application submitted successfully ({total_updates} fields filled).",
            }

        # --- Navigate to the next step ---
        next_clicked = await _smart_form_navigation(page)
        if next_clicked:
            progressed = True
            stall_count = 0
            await page.wait_for_timeout(1200)
            # After each Next, immediately clear any popup
            await _wait_and_dismiss_save_popup(page, timeout_ms=2000)
            await _fill_phone_if_possible(page, profile.get("phone", phone))
            await _fill_cover_letter_if_possible(page, profile.get("cover_letter", cover_letter))
            total_updates += await _enhanced_auto_fill_application_fields(page, profile, resume_path)
            continue

        stall_count += 1
        if stall_count >= 3:
            break
        await page.wait_for_timeout(1200)

    # Leave dialog open so user can finish manually
    return {
        "ok": False,
        "job_url": job_url,
        "status": "manual_review_required",
        "progressed": progressed,
        "field_updates": total_updates,
        "message": (
            f"Could not reach submit step automatically ({total_updates} fields filled). "
            "The application dialog may still be open — complete it manually in the browser."
        ),
    }


def _ensure_excel(file_path: Path) -> None:
    if file_path.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    ws.append(["Applied At (UTC)", "Title", "Company", "Location", "URL", "Status", "Notes"])
    wb.save(file_path)


def _append_jobs_to_excel(jobs: List[Dict[str, Any]], file_path: Path) -> Dict[str, Any]:
    _ensure_excel(file_path)
    wb = load_workbook(file_path)
    ws = wb["Applications"]

    existing_urls = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 5 and row[4]:
            existing_urls.add(str(row[4]).strip())

    added = 0
    skipped = 0
    for job in jobs:
        url = str(job.get("url", "")).strip()
        if not url or url in existing_urls:
            skipped += 1
            continue
        ws.append(
            [
                datetime.utcnow().isoformat() + "Z",
                job.get("title", ""),
                job.get("company", ""),
                job.get("location", ""),
                url,
                job.get("status", "unknown"),
                job.get("notes", ""),
            ]
        )
        existing_urls.add(url)
        added += 1

    wb.save(file_path)
    return {"added": added, "skipped": skipped, "file": str(file_path.resolve())}


@server.list_tools()
async def list_tools() -> List[types.Tool]:
    return [
        types.Tool(
            name="search_jobs",
            description="Search LinkedIn Easy Apply jobs by role and location (last 24 hours).",
            inputSchema={
                "type": "object",
                "properties": {
                    "role": {"type": "string", "description": "Role title, e.g. Python Developer"},
                    "location": {"type": "string", "description": "City or region, e.g. Bangalore"},
                    "count": {"type": "integer", "minimum": 1, "maximum": 100, "default": 10},
                    "user_data_dir": {"type": "string"},
                },
                "required": ["role", "location"],
            },
        ),
        types.Tool(
            name="easy_apply",
            description=(
                "Apply to a LinkedIn job URL using Easy Apply. "
                # FIX 2: Clearly document that dry_run defaults to False now
                "Set dry_run=false to actually submit (default is false)."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "job_url": {"type": "string"},
                    "cover_letter": {"type": "string", "default": ""},
                    "phone": {"type": "string", "default": ""},
                    "resume_path": {"type": "string", "default": ""},
                    "candidate_profile": {"type": "object"},
                    # FIX 2: Changed default from True to False
                    "dry_run": {"type": "boolean", "default": False},
                    "user_data_dir": {"type": "string"},
                },
                "required": ["job_url"],
            },
        ),
        types.Tool(
            name="enhanced_easy_apply",
            description=(
                "Enhanced LinkedIn job application with improved form handling. "
                "Set dry_run=false to actually submit (default is false)."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "job_url": {"type": "string"},
                    "cover_letter": {"type": "string", "default": ""},
                    "phone": {"type": "string", "default": ""},
                    "resume_path": {"type": "string", "default": ""},
                    "candidate_profile": {"type": "object"},
                    # FIX 2: Changed default from True to False
                    "dry_run": {"type": "boolean", "default": False},
                    "user_data_dir": {"type": "string"},
                },
                "required": ["job_url"],
            },
        ),
        types.Tool(
            name="save_to_excel",
            description="Save job application records to an Excel file with URL deduplication.",
            inputSchema={
                "type": "object",
                "properties": {
                    "jobs": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "title": {"type": "string"},
                                "company": {"type": "string"},
                                "location": {"type": "string"},
                                "url": {"type": "string"},
                                "status": {"type": "string"},
                                "notes": {"type": "string"},
                            },
                            "required": ["title", "company", "location", "url"],
                        },
                    },
                    "file_path": {"type": "string", "default": DEFAULT_EXCEL_FILE},
                },
                "required": ["jobs"],
            },
        ),
        types.Tool(
            name="debug_job_page",
            description="Inspect a LinkedIn job page and return visible button/link diagnostics.",
            inputSchema={
                "type": "object",
                "properties": {
                    "job_url": {"type": "string"},
                    "user_data_dir": {"type": "string"},
                },
                "required": ["job_url"],
            },
        ),
    ]


@server.call_tool()
async def call_tool(name: str, arguments: Dict[str, Any]) -> List[types.TextContent]:

    # ── search_jobs ──────────────────────────────────────────────────────────
    if name == "search_jobs":
        role = _clean_text(arguments.get("role"))
        location = _clean_text(arguments.get("location"))
        count = max(1, min(_safe_int(arguments.get("count"), 10), 100))
        user_data_dir = arguments.get("user_data_dir")

        if not role or not location:
            return _error("Both 'role' and 'location' are required.")

        context = None
        try:
            context = await _launch_context(user_data_dir)
            page = context.pages[0] if context.pages else await context.new_page()

            logged_in = await _ensure_logged_in(page)
            if not logged_in:
                profile_path = str(
                    (Path(user_data_dir) if user_data_dir else Path(DEFAULT_USER_DATA_DIR)).resolve()
                )
                return _error(
                    "LinkedIn login required. "
                    "Run the one-time login helper script to save your session, then retry.",
                    {"profile_dir": profile_path},
                )

            jobs = await _linkedin_search(page, role, location, count)
            return _ok({
                "ok": True,
                "tool": "search_jobs",
                "query": {"role": role, "location": location, "count": count},
                "result_count": len(jobs),
                "jobs": [job.to_dict() for job in jobs],
            })
        except PlaywrightTimeoutError:
            return _error("Timed out while searching LinkedIn jobs.")
        except Exception as exc:
            return _error("Unexpected error in search_jobs.", {"exception": str(exc)})
        finally:
            if context:
                await _close_context(context)

    # ── easy_apply & enhanced_easy_apply ─────────────────────────────────────
    if name in ("easy_apply", "enhanced_easy_apply"):
        job_url = _normalize_url(str(arguments.get("job_url", "")).strip())
        cover_letter = str(arguments.get("cover_letter", "")).strip()
        phone = str(arguments.get("phone", "")).strip()
        resume_path = str(arguments.get("resume_path", "")).strip()
        candidate_profile = arguments.get("candidate_profile")
        # FIX 2: Default is now False so it actually submits unless told otherwise
        dry_run = bool(arguments.get("dry_run", False))
        user_data_dir = arguments.get("user_data_dir")

        if not job_url:
            return _error("'job_url' is required.")

        context = None
        try:
            context = await _launch_context(user_data_dir)
            page = context.pages[0] if context.pages else await context.new_page()

            logged_in = await _ensure_logged_in(page)
            if not logged_in:
                profile_path = str(
                    (Path(user_data_dir) if user_data_dir else Path(DEFAULT_USER_DATA_DIR)).resolve()
                )
                return _error(
                    "LinkedIn login required. "
                    "Run the one-time login helper script to save your session, then retry.",
                    {"profile_dir": profile_path},
                )

            profile = _build_candidate_profile(
                phone=phone,
                cover_letter=cover_letter,
                resume_path=resume_path,
                candidate_profile=candidate_profile if isinstance(candidate_profile, dict) else None,
            )

            result = await _run_easy_apply_flow(
                page=page,
                job_url=job_url,
                profile=profile,
                phone=phone,
                cover_letter=cover_letter,
                dry_run=dry_run,
                resume_path=resume_path,
                enhanced=(name == "enhanced_easy_apply"),
            )
            result["tool"] = name
            # FIX 3: Always wrap result in _ok() — never return a raw dict
            return _ok(result)

        except PlaywrightTimeoutError:
            return _error(f"Timed out while performing {name}.")
        except Exception as exc:
            return _error(f"Unexpected error in {name}.", {"exception": str(exc)})
        finally:
            if context:
                await _close_context(context)

    # ── save_to_excel ─────────────────────────────────────────────────────────
    if name == "save_to_excel":
        jobs = arguments.get("jobs", [])
        file_path = Path(arguments.get("file_path", DEFAULT_EXCEL_FILE))

        if not isinstance(jobs, list) or not jobs:
            return _error("'jobs' must be a non-empty array.")

        try:
            output = _append_jobs_to_excel(jobs, file_path)
            return _ok({
                "ok": True,
                "tool": "save_to_excel",
                "file": output["file"],
                "added": output["added"],
                "skipped": output["skipped"],
            })
        except Exception as exc:
            return _error("Unexpected error in save_to_excel.", {"exception": str(exc)})

    # ── debug_job_page ────────────────────────────────────────────────────────
    if name == "debug_job_page":
        job_url = _normalize_url(str(arguments.get("job_url", "")).strip())
        user_data_dir = arguments.get("user_data_dir")

        if not job_url:
            return _error("'job_url' is required.")

        context = None
        try:
            context = await _launch_context(user_data_dir)
            page = context.pages[0] if context.pages else await context.new_page()

            logged_in = await _ensure_logged_in(page)
            if not logged_in:
                return _error("LinkedIn login required.")

            result = await _debug_job_page(page, job_url)
            return _ok(result)
        except PlaywrightTimeoutError:
            return _error("Timed out while inspecting job page.")
        except Exception as exc:
            return _error("Unexpected error in debug_job_page.", {"exception": str(exc)})
        finally:
            if context:
                await _close_context(context)

    return _error("Unknown tool.", {"tool": name})


async def run() -> None:
    async with mcp.server.stdio.stdio_server() as (read_stream, write_stream):
        await server.run(
            read_stream,
            write_stream,
            InitializationOptions(
                server_name=APP_NAME,
                server_version=APP_VERSION,
                capabilities=server.get_capabilities(
                    notification_options=NotificationOptions(),
                    experimental_capabilities={},
                ),
            ),
        )


if __name__ == "__main__":
    asyncio.run(run())